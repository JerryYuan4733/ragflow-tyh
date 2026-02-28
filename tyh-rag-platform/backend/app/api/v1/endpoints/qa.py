"""
Q&A管理接口
list, create, update, delete, import, template, versions, sync-to-ragflow, sync-from-ragflow

QA同步策略 V3:
- 正向同步: 分组路由推送 + 已修改QA更新 + 推送后回写所属KB
- 反向同步: 从 RAGFlow 知识库拉取 QA chunks 导入管理系统
- 同步覆盖: ragflow_sync QA 仅在 is_modified=true 时可推送
"""

import io
import uuid
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user, require_kb_admin
from app.db.session import get_db
from app.models import User, QAMeta, QAStatus, QASource, TeamDataset
from app.adapters.ragflow_client import ragflow_client
from app.services.team_service import TeamService
from app.services.qa_duplicate_detector import QADuplicateDetector
from pydantic import BaseModel

router = APIRouter()
logger = logging.getLogger(__name__)

# ==================== 正向同步常量 (V3) ====================
QA_SYNC_MAX_PER_FILE = 1000              # 单个 XLSX 文件 QA 上限
QA_SYNC_MAX_APPEND_TOTAL = 1000          # 追加模式下文档 chunk 上限
QA_SYNC_FILENAME_PREFIX = "qa_sync"      # 推送文件名前缀
QA_SYNC_CHUNK_TEMPLATE = "Question: {q}\tAnswer: {a}"  # 追加 chunk 内容格式
DATASET_FILTER_NONE = "__none__"                      # FR-36: 前端传此值表示筛选无归属知识库的 QA

class QACreateRequest(BaseModel):
    question: str
    answer: str


class QAUpdateRequest(BaseModel):
    question: Optional[str] = None
    answer: Optional[str] = None


# ==================== 正向同步 V3 核心逻辑 ====================


async def _sync_single_dataset(
    dataset_id: str,
    qa_list: list,
    all_active_questions: set[str],
) -> dict:
    """
    正向同步 V3 单知识库推送：将指定 QA 列表推送到 RAGFlow 指定知识库。

    沿用 V2 策略:
    - C-18: 增量同步 — 清理 RAGFlow 中多余的 chunk
    - 策略 A: 追加 chunk（chunks + 待推送 ≤ 1000）
    - 策略 B: 生成 XLSX 上传
    V3 新增:
    - 分离 normal/modified QA
    - 对 modified QA 执行"删旧 chunk + 追加新 chunk"（FR-33）
    - 返回 updated 计数
    """
    from datetime import datetime

    # 1. 分离普通 QA 和已修改 QA
    normal_qas = []
    modified_qas = []
    for qa in qa_list:
        if getattr(qa, "is_modified", False):
            modified_qas.append(qa)
        else:
            normal_qas.append(qa)

    normal_pairs = [(qa.question, qa.answer) for qa in normal_qas]
    modified_questions = {qa.question for qa in modified_qas}

    logger.info(
        f"V3 单KB推送: dataset={dataset_id}, "
        f"普通={len(normal_pairs)}, 已修改={len(modified_qas)}"
    )

    # 2. 获取目标知识库中 QA 文档
    qa_docs = await ragflow_client.list_qa_documents(dataset_id)
    logger.info(f"V3 单KB推送: QA 文档数={len(qa_docs)}")

    # 3. 确定目标文档（chunk 最多的）
    target_doc = None
    target_chunk_count = 0
    if qa_docs:
        target_doc = max(qa_docs, key=lambda d: d.chunk_count)
        target_chunk_count = await ragflow_client.get_document_chunk_count(
            dataset_id, target_doc.id
        )
        logger.info(
            f"V3 单KB推送: 候选追加文档={target_doc.name}, "
            f"chunk_count={target_doc.chunk_count}, 精确count={target_chunk_count}"
        )

    # 4. C-18: 清理不活跃 chunk + 收集 question 位置映射
    cleaned, question_locations = await _cleanup_inactive_chunks(
        dataset_id, qa_docs, all_active_questions
    )

    # 5. FR-33: 对已修改 QA 执行"删旧 + 追新"
    updated = 0
    if modified_qas:
        if target_doc:
            updated = await _update_modified_chunks(
                dataset_id, target_doc.id, modified_qas, question_locations
            )
        else:
            # 无 QA 文档 → 合并到普通推送（无旧 chunk 可删）
            normal_pairs.extend([(qa.question, qa.answer) for qa in modified_qas])
            logger.info(f"V3: 无QA文档，{len(modified_qas)} 条已修改QA合并到普通推送")

    # 6. 无待推送的普通 QA 时提前返回
    if not normal_pairs and not modified_qas:
        msg = "无待推送的 QA"
        if cleaned > 0:
            msg += f"，已清理 {cleaned} 个多余 chunk"
        return {
            "strategy": "cleanup_only" if cleaned > 0 else "none",
            "message": msg,
            "appended": 0, "skipped": 0, "updated": updated,
            "cleaned": cleaned, "uploaded_files": 0, "total_qa": 0,
            "file_names": [],
        }

    if not normal_pairs:
        # 只有已修改 QA，无普通 QA 需要追加/上传
        msg = f"更新 {updated} 条已修改QA"
        if cleaned > 0:
            msg += f"，清理 {cleaned} 个多余 chunk"
        return {
            "strategy": "update_only",
            "message": msg,
            "appended": 0, "skipped": 0, "updated": updated,
            "cleaned": cleaned, "uploaded_files": 0, "total_qa": updated,
            "file_names": [],
        }

    # 7. 策略选择（仅针对普通 QA）
    can_append = (
        target_doc is not None
        and (target_chunk_count + len(normal_pairs)) <= QA_SYNC_MAX_APPEND_TOTAL
    )

    appended_questions: set[str] = set()
    if can_append:
        push_result = await _strategy_append_chunks(
            dataset_id, target_doc.id, target_doc.name, normal_pairs
        )
        appended_questions = push_result.get("_appended_questions", set())
    else:
        push_result = await _strategy_upload_xlsx(
            dataset_id, normal_pairs, datetime.now()
        )

    # 8. C-23: 跨文档去重（已修改 QA 的 question 也算"刚推送"）
    target_doc_id = target_doc.id if target_doc else ""
    all_pushed_questions = appended_questions | modified_questions
    deduped = _dedup_cross_doc(
        question_locations, target_doc_id, all_pushed_questions
    )
    if deduped:
        deleted = await _execute_dedup_deletes(dataset_id, deduped)
        cleaned += deleted

    push_result.pop("_appended_questions", None)
    push_result["updated"] = updated
    push_result["cleaned"] = cleaned
    if updated > 0:
        push_result["message"] += f"，更新 {updated} 条已修改QA"
    if cleaned > 0:
        push_result["message"] += f"，清理 {cleaned} 个多余 chunk"
    return push_result


async def _cleanup_inactive_chunks(
    dataset_id: str,
    qa_docs: list,
    all_active_questions: set[str],
) -> tuple[int, dict[str, list[tuple[str, str, str]]]]:
    """
    C-18: 扫描 RAGFlow QA 文档，删除不对应任何活跃 QA 的多余 chunk。
    同时收集活跃 chunk 的 question → [(doc_id, doc_name, chunk_id)] 映射，
    供 C-23 跨文档去重使用（避免追加后重新扫描的最终一致性问题）。
    返回 (删除数, question_locations)。
    """
    cleaned = 0
    question_locations: dict[str, list[tuple[str, str, str]]] = {}
    for doc in qa_docs:
        try:
            chunks = await ragflow_client.list_all_chunks(dataset_id, doc.id)
            for chunk in chunks:
                q, _ = _extract_qa_from_chunk(chunk)
                chunk_id = chunk.get("id", "")
                if not chunk_id:
                    continue
                if q and q not in all_active_questions:
                    try:
                        await ragflow_client.delete_chunk(dataset_id, doc.id, chunk_id)
                        cleaned += 1
                    except Exception as e:
                        logger.warning(f"C-18 清理 chunk 失败: doc={doc.name}, chunk={chunk_id}: {e}")
                elif q:
                    # 记录活跃 chunk 的位置（C-23 去重用）
                    question_locations.setdefault(q, []).append(
                        (doc.id, doc.name, chunk_id)
                    )
        except Exception as e:
            logger.warning(f"C-18 清理文档 {doc.name} 时出错: {e}")
    if cleaned > 0:
        logger.info(f"C-18 清理: 删除 {cleaned} 个不活跃 chunk")
    return cleaned, question_locations


def _dedup_cross_doc(
    question_locations: dict[str, list[tuple[str, str, str]]],
    target_doc_id: str,
    appended_questions: set[str],
) -> list[tuple[str, str, str]]:
    """
    C-23: 纯计算 — 根据 C-18 扫描数据和追加结果，确定需要删除的跨文档重复 chunk。
    判断逻辑: question 在目标文档中已有(in_target)或刚被追加(in appended_questions)
              → 删除非目标文档中的同名 chunk。
    返回待删除列表 [(doc_id, doc_name, chunk_id), ...]。
    """
    if not target_doc_id:
        return []
    to_delete: list[tuple[str, str, str]] = []
    for q, locations in question_locations.items():
        # 判断目标文档中是否已有，或刚被追加
        in_target = any(doc_id == target_doc_id for doc_id, _, _ in locations)
        just_appended = q in appended_questions
        if not in_target and not just_appended:
            continue  # 目标文档中没有且没追加，保持现状
        for doc_id, doc_name, chunk_id in locations:
            if doc_id == target_doc_id:
                continue  # 保留目标文档中的副本
            to_delete.append((doc_id, doc_name, chunk_id))
    if to_delete:
        logger.info(
            f"C-23 跨文档去重: 计划删除 {len(to_delete)} 个重复 chunk "
            f"(in_target + appended={len(appended_questions)})"
        )
    return to_delete


async def _execute_dedup_deletes(
    dataset_id: str,
    to_delete: list[tuple[str, str, str]],
) -> int:
    """C-23: 执行跨文档去重删除。"""
    deleted = 0
    for doc_id, doc_name, chunk_id in to_delete:
        try:
            await ragflow_client.delete_chunk(dataset_id, doc_id, chunk_id)
            deleted += 1
            logger.info(f"C-23 已删除: doc={doc_name}, chunk={chunk_id}")
        except Exception as e:
            logger.warning(f"C-23 跨文档去重失败: doc={doc_name}, chunk={chunk_id}: {e}")
    if deleted > 0:
        logger.info(f"C-23 跨文档去重: 共删除 {deleted} 个重复 chunk")
    return deleted


async def _update_modified_chunks(
    dataset_id: str,
    target_doc_id: str,
    modified_qas: list,
    question_locations: dict[str, list[tuple[str, str, str]]],
) -> int:
    """
    FR-33: 对已修改的 QA（is_modified=true）执行"删旧 chunk + 追加新 chunk"。
    1. 按 previous_question（如有）在 question_locations 中查找旧 chunk → 删除
    2. 按当前 question 在 question_locations 中查找旧 chunk → 删除
    3. 追加新 chunk 到目标文档
    返回成功更新数。
    """
    from app.adapters.ragflow_types import ChunkCreate

    updated = 0
    for qa in modified_qas:
        try:
            # 步骤1: 删除 previous_question 对应的旧 chunk（question 被改过的场景）
            if qa.previous_question and qa.previous_question in question_locations:
                for doc_id, doc_name, chunk_id in question_locations[qa.previous_question]:
                    try:
                        await ragflow_client.delete_chunk(dataset_id, doc_id, chunk_id)
                        logger.info(
                            f"FR-33 更新: 删除旧question chunk "
                            f"doc={doc_name}, q='{qa.previous_question[:50]}'"
                        )
                    except Exception as e:
                        logger.warning(f"FR-33 删除旧question chunk失败: {e}")

            # 步骤2: 删除当前 question 对应的旧 chunk（answer 被改的场景）
            if qa.question in question_locations:
                for doc_id, doc_name, chunk_id in question_locations[qa.question]:
                    try:
                        await ragflow_client.delete_chunk(dataset_id, doc_id, chunk_id)
                        logger.info(
                            f"FR-33 更新: 删除同名question chunk "
                            f"doc={doc_name}, q='{qa.question[:50]}'"
                        )
                    except Exception as e:
                        logger.warning(f"FR-33 删除同名question chunk失败: {e}")

            # 步骤3: 追加新 chunk 到目标文档
            content = QA_SYNC_CHUNK_TEMPLATE.format(q=qa.question, a=qa.answer)
            await ragflow_client.create_chunk(
                dataset_id, target_doc_id, ChunkCreate(content=content)
            )
            updated += 1
            logger.info(f"FR-33 更新: 追加新chunk q='{qa.question[:50]}'")
        except Exception as e:
            logger.error(f"FR-33 更新QA失败: q='{qa.question[:50]}': {e}")

    if updated > 0:
        logger.info(f"FR-33: 共更新 {updated} 条已修改QA")
    return updated


async def _push_router(
    db: AsyncSession,
    target_dataset_id: Optional[str],
    qa_list: list,
    all_active_questions: set[str],
    team_id: str,
) -> list[dict]:
    """
    FR-34: 推送路由器 — 按 ragflow_dataset_id 分组，对每组调用 _sync_single_dataset。
    - 有所属 KB 的 QA → 推到各自的 KB
    - 无所属 KB 的 QA → 推到 target_dataset_id
    推送成功后回写 QAMeta.ragflow_dataset_id 并重置 is_modified/previous_question。
    """
    # 1. 按 ragflow_dataset_id 分组
    groups: dict[str, list] = {}
    unassigned: list = []
    for qa in qa_list:
        if qa.ragflow_dataset_id:
            groups.setdefault(qa.ragflow_dataset_id, []).append(qa)
        else:
            unassigned.append(qa)

    # 2. 无归属 QA 分配到目标 KB
    if unassigned:
        if not target_dataset_id:
            raise HTTPException(
                status_code=400,
                detail="存在无所属知识库的QA，请选择目标知识库"
            )
        groups.setdefault(target_dataset_id, []).extend(unassigned)

    # 3. 预加载知识库名称（T-2.9）
    all_dataset_ids = list(groups.keys())
    name_map: dict[str, str] = {}
    if all_dataset_ids:
        name_result = await db.execute(
            select(
                TeamDataset.ragflow_dataset_id,
                TeamDataset.ragflow_dataset_name,
            ).where(TeamDataset.ragflow_dataset_id.in_(all_dataset_ids))
        )
        for row in name_result.all():
            name_map[row[0]] = row[1]

    # 4. 对每个分组执行推送
    results = []
    for ds_id, group_qas in groups.items():
        dataset_name = name_map.get(ds_id, ds_id)
        try:
            result = await _sync_single_dataset(
                ds_id, group_qas, all_active_questions
            )
            result["dataset_id"] = ds_id
            result["dataset_name"] = dataset_name

            # FR-32: 推送成功后回写知识库归属 + 重置修改标记（T-2.8）
            for qa in group_qas:
                qa.ragflow_dataset_id = ds_id
                qa.is_modified = False
                qa.previous_question = None
            await db.flush()

            results.append(result)
            logger.info(f"FR-34 分组推送完成: {dataset_name}")
        except HTTPException:
            raise
        except Exception as e:
            logger.error(
                f"FR-34 分组推送失败: dataset={dataset_name}: {e}",
                exc_info=True,
            )
            results.append({
                "dataset_id": ds_id,
                "dataset_name": dataset_name,
                "strategy": "error",
                "message": f"推送失败: {str(e)}",
                "appended": 0, "skipped": 0, "updated": 0,
                "cleaned": 0, "uploaded_files": 0, "total_qa": 0,
                "file_names": [],
            })

    return results


async def _strategy_append_chunks(
    dataset_id: str,
    document_id: str,
    document_name: str,
    qa_pairs: list[tuple[str, str]],
) -> dict:
    """策略 A: 追加 chunk 到已有 QA 文档，跳过已存在的 question"""
    from app.adapters.ragflow_types import ChunkCreate

    # 获取已有 chunk 的 question 集合（去重用）
    existing_chunks = await ragflow_client.list_all_chunks(dataset_id, document_id)
    existing_questions: set[str] = set()
    for chunk in existing_chunks:
        q, _ = _extract_qa_from_chunk(chunk)
        if q:
            existing_questions.add(q)
    original_existing = set(existing_questions)  # 快照，用于计算追加差集
    logger.info(f"策略A: 文档 {document_name} 已有 {len(existing_questions)} 个唯一 question")

    appended = 0
    skipped = 0
    for question, answer in qa_pairs:
        if question in existing_questions:
            skipped += 1
            continue
        content = QA_SYNC_CHUNK_TEMPLATE.format(q=question, a=answer)
        try:
            await ragflow_client.create_chunk(
                dataset_id, document_id, ChunkCreate(content=content)
            )
            existing_questions.add(question)
            appended += 1
        except Exception as e:
            logger.warning(f"策略A: 追加 chunk 失败: {e}")
            # 降级: 剩余 QA 走策略 B
            from datetime import datetime
            remaining = [(q, a) for q, a in qa_pairs if q not in existing_questions]
            if remaining:
                logger.info(f"策略A 降级: {len(remaining)} 条 QA 改用策略B上传")
                fallback = await _strategy_upload_xlsx(
                    dataset_id, remaining, datetime.now()
                )
                return {
                    "strategy": "append+fallback",
                    "message": f"追加 {appended} 条后降级上传 {fallback['total_qa']} 条",
                    "appended": appended,
                    "skipped": skipped,
                    "uploaded_files": fallback["uploaded_files"],
                    "total_qa": appended + fallback["total_qa"],
                    "file_names": fallback["file_names"],
                }
            break

    # 收集实际追加的 question（供 C-23 跨文档去重判断）
    appended_set = existing_questions - original_existing
    msg = f"追加 {appended} 条到文档 {document_name}"
    if skipped > 0:
        msg += f"，跳过 {skipped} 条已存在"
    return {
        "strategy": "append",
        "message": msg,
        "appended": appended,
        "skipped": skipped,
        "uploaded_files": 0,
        "total_qa": appended,
        "file_names": [],
        "_appended_questions": appended_set,
    }


async def _strategy_upload_xlsx(
    dataset_id: str,
    qa_pairs: list[tuple[str, str]],
    now,
) -> dict:
    """策略 B: 生成 XLSX 无标题两列并上传到 RAGFlow"""
    files = _build_qa_xlsx(qa_pairs, now)
    uploaded_names = []

    for filename, content_bytes in files:
        try:
            doc_ids = await ragflow_client.upload_documents(
                dataset_id,
                [(filename, content_bytes,
                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")],
            )
            if doc_ids:
                doc_id = doc_ids[0]
                await ragflow_client.update_document_parser(dataset_id, doc_id, "qa")
                await ragflow_client.start_parsing(dataset_id, [doc_id])
                uploaded_names.append(filename)
                logger.info(f"策略B: 上传成功 {filename}, doc_id={doc_id}")
            else:
                logger.error(f"策略B: 上传 {filename} 返回空 doc_ids")
        except Exception as e:
            logger.error(f"策略B: 上传 {filename} 失败: {e}", exc_info=True)

    msg = f"上传 {len(uploaded_names)} 个文件，共 {len(qa_pairs)} 条 QA"
    return {
        "strategy": "upload",
        "message": msg,
        "appended": 0,
        "skipped": 0,
        "uploaded_files": len(uploaded_names),
        "total_qa": len(qa_pairs),
        "file_names": uploaded_names,
    }


def _build_qa_xlsx(
    qa_pairs: list[tuple[str, str]],
    now=None,
    max_per_file: int = QA_SYNC_MAX_PER_FILE,
) -> list[tuple[str, bytes]]:
    """
    生成 RAGFlow QA 规范的 XLSX 文件（无标题行，两列: 问题, 答案）。
    单文件上限 max_per_file 条，超出分多个文件。
    返回 [(文件名, 字节内容), ...]
    """
    import openpyxl
    from datetime import datetime

    if now is None:
        now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    files: list[tuple[str, bytes]] = []
    total_batches = (len(qa_pairs) + max_per_file - 1) // max_per_file

    for i in range(0, len(qa_pairs), max_per_file):
        batch = qa_pairs[i:i + max_per_file]
        wb = openpyxl.Workbook()
        ws = wb.active
        for q, a in batch:
            ws.append([q, a])

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        file_idx = i // max_per_file + 1
        suffix = f"_{file_idx}" if total_batches > 1 else ""
        filename = f"{QA_SYNC_FILENAME_PREFIX}_{timestamp}{suffix}.xlsx"
        files.append((filename, buf.getvalue()))

    logger.info(f"XLSX 生成: {len(files)} 个文件, 共 {len(qa_pairs)} 条 QA")
    return files


async def _sync_from_ragflow(
    db: AsyncSession, dataset_id: str, team_id: str, user_id: str,
) -> dict:
    """
    反向同步 V2：从 RAGFlow 知识库差异同步 QA 到管理系统。

    差异同步策略:
    - 新增: RAGFlow 有但本地没有 → 插入 (source=ragflow_sync)
    - 更新: 双方 question 相同但 answer 不同 → 以 RAGFlow 为准覆盖 (C-17)
    - 删除: 本地 source=ragflow_sync 且 ragflow_dataset_id 匹配，但 RAGFlow 中已不存在 → 硬删除 (C-16)
    """
    imported = 0
    updated = 0
    deleted = 0
    skipped = 0
    parse_failed = 0
    errors = 0

    try:
        # 1. 从 RAGFlow 构建 question→answer 映射（仅 parser_id=qa 文档）
        ragflow_qa_map: dict[str, str] = {}
        all_documents = await ragflow_client.list_documents(dataset_id, page=1, size=100)
        qa_documents = [d for d in all_documents if d.effective_parser == "qa"]
        logger.info(
            f"反向同步: dataset={dataset_id}, "
            f"总文档={len(all_documents)}, QA文档={len(qa_documents)}"
        )

        for doc in qa_documents:
            try:
                chunks = await ragflow_client.list_all_chunks(dataset_id, doc.id)
                logger.info(f"文档 {doc.name}: {len(chunks)} 个 chunks")
                if chunks:
                    sample = chunks[0]
                    logger.info(
                        f"文档 {doc.name} 示例 chunk: keys={list(sample.keys())}, "
                        f"content[:120]={repr(str(sample.get('content', ''))[:120])}"
                    )
                for chunk in chunks:
                    question, answer = _extract_qa_from_chunk(chunk)
                    if not question or not answer:
                        parse_failed += 1
                        continue
                    ragflow_qa_map[question] = answer  # 同 question 以最后一条为准
            except Exception as e:
                logger.warning(f"处理文档 {doc.name} 时出错: {e}", exc_info=True)
                errors += 1

        logger.info(f"反向同步: RAGFlow 有效 QA 数={len(ragflow_qa_map)}")

        # 2. 加载本地该团队所有 QA（含所有 source）
        result = await db.execute(
            select(QAMeta).where(QAMeta.team_id == team_id)
        )
        local_qas: dict[str, QAMeta] = {}
        for qa in result.scalars().all():
            local_qas[qa.question] = qa

        # 3. 新增 + 更新: 遍历 RAGFlow QA
        for question, answer in ragflow_qa_map.items():
            if question in local_qas:
                local_qa = local_qas[question]
                if local_qa.answer != answer:
                    # C-17: 覆盖更新 — 以 RAGFlow 为准
                    local_qa.answer = answer
                    local_qa.answer_summary = answer[:200]
                    local_qa.version += 1
                    local_qa.edited_by = user_id
                    updated += 1
                else:
                    skipped += 1
            else:
                # 新增 QA
                qa = QAMeta(
                    id=str(uuid.uuid4()),
                    team_id=team_id,
                    question=question,
                    answer=answer,
                    question_summary=question[:200],
                    answer_summary=answer[:200],
                    source="ragflow_sync",
                    ragflow_dataset_id=dataset_id,
                    edited_by=user_id,
                )
                db.add(qa)
                imported += 1

        # 4. C-16: 删除本地已不存在于 RAGFlow 的 QA（仅 source=ragflow_sync + 同 dataset）
        for question, qa in local_qas.items():
            if (
                qa.source == QASource.RAGFLOW_SYNC
                and qa.ragflow_dataset_id == dataset_id
                and question not in ragflow_qa_map
            ):
                await db.delete(qa)
                deleted += 1

        await db.flush()

        logger.info(
            f"反向同步结果: 新增={imported}, 更新={updated}, "
            f"删除={deleted}, 跳过={skipped}, 解析失败={parse_failed}, 错误={errors}"
        )

    except Exception as e:
        logger.error(f"反向同步失败: {e}", exc_info=True)
        raise

    return {
        "imported": imported, "updated": updated, "deleted": deleted,
        "skipped": skipped, "parse_failed": parse_failed, "errors": errors,
    }


def _extract_qa_from_chunk(chunk: dict) -> tuple[str, str]:
    """
    从 RAGFlow chunk dict 中提取 question 和 answer。
    经诊断，RAGFlow API 返回的 chunk 字段:
      available, content, dataset_id, docnm_kwd, document_id,
      id, image_id, important_keywords, positions, questions
    注意: API 响应中**没有 content_with_weight 字段**，QA 内容全部在 content 中。

    QA 解析模式的 content 格式:
      "Question: <问题>     Answer: <答案>"  （标签 + 空格分隔）
      "Question: <问题>\\nAnswer: <答案>"    （标签 + 换行分隔）
    """
    content = str(chunk.get("content", "")).strip()
    if not content:
        return "", ""

    # 策略1（最高优先级）: "Question: ... Answer: ..." 标签格式
    #   RAGFlow QA 解析模式的标准输出格式
    q, a = _parse_labeled_qa(content)
    if q and a:
        return q, a

    # 策略2: content_with_weight 和 content 分别存储 Q 和 A（兼容其他版本）
    content_with_weight = str(chunk.get("content_with_weight", "")).strip()
    if (content_with_weight and content
            and content_with_weight != content
            and '\n' not in content_with_weight
            and not content_with_weight.lower().startswith("question:")):
        return content_with_weight, content

    # 策略3: CSV 双引号 或 换行分割
    q, a = _parse_combined_qa(content)
    if q and a:
        return q, a

    return "", ""


def _parse_labeled_qa(content: str) -> tuple[str, str]:
    """
    解析 RAGFlow QA 解析模式的标签格式:
      "Question: <问题>  Answer: <答案>"
      "Question: <问题>\\nAnswer: <答案>"
    不区分大小写查找 'Answer:' 分隔符。
    """
    lower = content.lower()
    # 查找 "answer:" 分隔符（不区分大小写）
    answer_idx = lower.find("answer:")
    if answer_idx == -1:
        return "", ""

    q_part = content[:answer_idx].strip()
    a_part = content[answer_idx + len("answer:"):].strip()

    # 去除 "Question:" 前缀（不区分大小写）
    q_lower = q_part.lower()
    if q_lower.startswith("question:"):
        q_part = q_part[len("question:"):].strip()

    if not q_part or not a_part:
        return "", ""

    return q_part, a_part


def _parse_combined_qa(content: str) -> tuple[str, str]:
    """
    从合并字符串中解析 question 和 answer（兜底格式）。
    支持格式:
    - CSV 双引号: "问题","答案"
    - 换行分割: 问题\\n答案
    """
    content = content.strip()
    if not content:
        return "", ""

    # 格式1: CSV 双引号包裹
    if content.startswith('"') and '","' in content:
        parts = content.split('","', 1)
        q = parts[0].strip('"').replace('""', '"')
        a = parts[1].rstrip('"').replace('""', '"') if len(parts) > 1 else ""
        return q.strip(), a.strip()

    # 格式2: 按第一个换行分割
    if '\n' in content:
        lines = content.split('\n', 1)
        return lines[0].strip(), lines[1].strip()

    # 无法解析
    return "", ""


# ==================== API 端点 ====================

@router.get("")
async def list_qa(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    dataset_id: Optional[str] = None,
    status: Optional[str] = None,
    source: Optional[str] = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取Q&A列表（支持按知识库/状态/来源筛选）"""
    query = select(QAMeta).where(QAMeta.team_id == user.active_team_id)
    if keyword:
        query = query.where(QAMeta.question.contains(keyword) | QAMeta.answer.contains(keyword))
    # T-13.1 + FR-36: 按知识库筛选（支持 __none__ 筛选无归属 QA）
    if dataset_id:
        if dataset_id == DATASET_FILTER_NONE:
            query = query.where(QAMeta.ragflow_dataset_id.is_(None))
        else:
            query = query.where(QAMeta.ragflow_dataset_id == dataset_id)
    # T-13.2: 按状态筛选
    if status:
        query = query.where(QAMeta.status == QAStatus(status))
    # 按来源筛选
    if source:
        query = query.where(QAMeta.source == QASource(source))

    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar() or 0
    result = await db.execute(query.order_by(QAMeta.created_at.desc()).offset((page - 1) * page_size).limit(page_size))
    items = result.scalars().all()

    # FR-31: 构建 dataset_id → dataset_name 映射
    dataset_name_map: dict[str, str] = {}
    ds_ids = {q.ragflow_dataset_id for q in items if q.ragflow_dataset_id}
    if ds_ids:
        from app.models import TeamDataset
        ds_result = await db.execute(
            select(TeamDataset.ragflow_dataset_id, TeamDataset.ragflow_dataset_name)
            .where(TeamDataset.ragflow_dataset_id.in_(ds_ids))
        )
        for row in ds_result.all():
            dataset_name_map[row[0]] = row[1] or row[0]

    # T-13.3: 响应新增 status/source/ragflow_dataset_id/ragflow_dataset_name/updated_at
    return {
        "items": [
            {"id": q.id, "question": q.question, "answer": q.answer,
             "version": q.version, "created_at": q.created_at.isoformat(),
             "status": q.status.value if hasattr(q.status, 'value') else str(q.status),
             "source": q.source.value if hasattr(q.source, 'value') else str(q.source),
             "ragflow_dataset_id": q.ragflow_dataset_id,
             "ragflow_dataset_name": dataset_name_map.get(q.ragflow_dataset_id, "") if q.ragflow_dataset_id else "",
             "updated_at": q.updated_at.isoformat() if q.updated_at else None}
            for q in items
        ],
        "total": total,
    }


# T-13.4: QA 状态切换接口
class QAStatusRequest(BaseModel):
    status: str  # active / pending_review / disabled


@router.put("/{qa_id}/status")
async def change_qa_status(
    qa_id: str,
    request: QAStatusRequest,
    user: User = Depends(require_kb_admin),
    db: AsyncSession = Depends(get_db),
):
    """切换 QA 状态"""
    try:
        new_status = QAStatus(request.status)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"无效状态值: {request.status}")

    result = await db.execute(select(QAMeta).where(QAMeta.id == qa_id))
    qa = result.scalar_one_or_none()
    if not qa:
        raise HTTPException(status_code=404, detail="Q&A不存在")

    qa.status = new_status
    qa.edited_by = user.id
    await db.flush()
    return {"message": "状态更新成功", "status": new_status.value}


@router.post("", status_code=201)
async def create_qa(
    request: QACreateRequest,
    user: User = Depends(require_kb_admin),
    db: AsyncSession = Depends(get_db),
):
    """新增Q&A（含重复检测）"""
    # QA 重复检测
    duplicate = await QADuplicateDetector.check(
        db, request.question, user.active_team_id
    )
    if duplicate:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "问题已存在，与 QA 重复",
                "duplicate_qa_id": duplicate.qa_id,
                "duplicate_question": duplicate.question,
                "match_type": duplicate.match_type,
                "similarity": duplicate.similarity,
            }
        )

    qa = QAMeta(
        id=str(uuid.uuid4()),
        team_id=user.active_team_id,
        question=request.question,
        answer=request.answer,
        question_summary=request.question[:200],
        answer_summary=request.answer[:200],
        edited_by=user.id,
    )
    db.add(qa)
    await db.flush()

    return {"id": qa.id, "message": "创建成功"}


@router.put("/{qa_id}")
async def update_qa(
    qa_id: str,
    request: QAUpdateRequest,
    user: User = Depends(require_kb_admin),
    db: AsyncSession = Depends(get_db),
):
    """更新Q&A"""
    result = await db.execute(select(QAMeta).where(QAMeta.id == qa_id))
    qa = result.scalar_one_or_none()
    if not qa:
        raise HTTPException(status_code=404, detail="Q&A不存在")
    # FR-33/C-25: 编辑标记（用于 ragflow_sync QA 推送判断）
    if request.question is not None and request.question != qa.question:
        qa.previous_question = qa.question  # C-29: 记录旧 question
    if request.question is not None or request.answer is not None:
        qa.is_modified = True  # 任何实质编辑都标记

    if request.question is not None:
        qa.question = request.question
        qa.question_summary = request.question[:200]
    if request.answer is not None:
        qa.answer = request.answer
        qa.answer_summary = request.answer[:200]
    qa.version += 1
    qa.edited_by = user.id
    await db.flush()

    return {"message": "更新成功", "version": qa.version}


@router.delete("/{qa_id}")
async def delete_qa(
    qa_id: str,
    user: User = Depends(require_kb_admin),
    db: AsyncSession = Depends(get_db),
):
    """删除Q&A"""
    result = await db.execute(select(QAMeta).where(QAMeta.id == qa_id))
    qa = result.scalar_one_or_none()
    if not qa:
        raise HTTPException(status_code=404, detail="Q&A不存在")
    await db.delete(qa)
    await db.flush()

    return {"message": "删除成功"}


class SyncRequest(BaseModel):
    dataset_id: Optional[str] = None  # 目标知识库（无所属KB的QA推送到此；反向同步可选）
    qa_ids: Optional[list[str]] = None  # V3: 指定推送的QA ID列表（空=全量）


@router.post("/sync-to-ragflow")
async def sync_to_ragflow(
    request: SyncRequest = SyncRequest(),
    user: User = Depends(require_kb_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    正向同步 V3：将当前团队 QA 推送到 RAGFlow，支持分组路由和勾选推送。
    - FR-32: 推送后回写 ragflow_dataset_id
    - FR-33: ragflow_sync + is_modified=true 的 QA 可推送
    - FR-34: 按所属知识库分组路由
    """
    team_id = user.active_team_id

    # 1. 构建查询条件（T-2.2 + T-2.3）
    conditions = [
        QAMeta.team_id == team_id,
        QAMeta.status == QAStatus.ACTIVE,
        # V3: ragflow_sync + is_modified 也可推送（FR-33）
        or_(
            QAMeta.source != QASource.RAGFLOW_SYNC,
            QAMeta.is_modified == True,
        ),
    ]
    if request.qa_ids:
        conditions.append(QAMeta.id.in_(request.qa_ids))

    result = await db.execute(
        select(QAMeta).where(*conditions).order_by(QAMeta.created_at.asc())
    )
    qa_list = list(result.scalars().all())

    if not qa_list:
        return {
            "message": "无符合条件的待推送 QA",
            "groups": [],
            "total_appended": 0, "total_updated": 0,
            "total_skipped": 0, "total_cleaned": 0,
        }

    # 2. 构建团队全部 active QA question 集合（含 ragflow_sync，用于 C-18 清理判断）
    all_active_result = await db.execute(
        select(QAMeta.question).where(
            QAMeta.team_id == team_id,
            QAMeta.status == QAStatus.ACTIVE,
        )
    )
    all_active_questions: set[str] = {row[0] for row in all_active_result.all()}

    logger.info(
        f"正向同步 V3: 团队={team_id}, "
        f"待推送={len(qa_list)}, 全部 active={len(all_active_questions)}"
    )

    # 3. 分组路由 + 推送 + 回写（T-2.4 ~ T-2.9）
    try:
        groups = await _push_router(
            db, request.dataset_id, qa_list, all_active_questions, team_id
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"正向同步 V3 错误: {e}", exc_info=True)
        raise HTTPException(status_code=502, detail=f"正向同步失败: {str(e)}")

    # 4. 汇总分组结果（T-2.10）
    total_appended = sum(g.get("appended", 0) for g in groups)
    total_updated = sum(g.get("updated", 0) for g in groups)
    total_skipped = sum(g.get("skipped", 0) for g in groups)
    total_cleaned = sum(g.get("cleaned", 0) for g in groups)

    parts = []
    for g in groups:
        name = g.get("dataset_name", "")
        sub = []
        if g.get("appended", 0) > 0:
            sub.append(f"追加{g['appended']}条")
        if g.get("updated", 0) > 0:
            sub.append(f"更新{g['updated']}条")
        if sub:
            parts.append(f"{name} {','.join(sub)}")
    message = "推送完成"
    if parts:
        message += ": " + ", ".join(parts)

    return {
        "message": message,
        "groups": groups,
        "total_appended": total_appended,
        "total_updated": total_updated,
        "total_skipped": total_skipped,
        "total_cleaned": total_cleaned,
    }


@router.post("/sync-from-ragflow")
async def sync_from_ragflow(
    request: SyncRequest = SyncRequest(),
    user: User = Depends(require_kb_admin),
    db: AsyncSession = Depends(get_db),
):
    """
    反向同步 V2：从 RAGFlow 知识库差异同步 QA 到管理系统。
    - 指定 dataset_id → 仅同步该知识库
    - 未指定 → 遍历团队全部知识库（BUG-10 修复）
    - 差异同步: 新增 + 更新答案 + 删除已不存在的（BUG-12 修复）
    """
    team_id = user.active_team_id

    # 确定要同步的知识库列表
    if request.dataset_id:
        dataset_ids = [request.dataset_id]
    else:
        dataset_ids = await TeamService.get_team_dataset_ids(db, team_id)
    if not dataset_ids:
        raise HTTPException(status_code=400, detail="团队未绑定知识库，无法反向同步")

    # 逐个知识库同步并汇总结果
    total = {"imported": 0, "updated": 0, "deleted": 0, "skipped": 0, "parse_failed": 0, "errors": 0}
    try:
        for ds_id in dataset_ids:
            logger.info(f"反向同步: 开始处理知识库 {ds_id}")
            result = await _sync_from_ragflow(db, ds_id, team_id, user.id)
            for key in total:
                total[key] += result.get(key, 0)

        parts = [f"反向同步完成（{len(dataset_ids)} 个知识库）"]
        if total['imported'] > 0:
            parts.append(f"新增 {total['imported']} 条")
        if total['updated'] > 0:
            parts.append(f"更新 {total['updated']} 条")
        if total['deleted'] > 0:
            parts.append(f"删除 {total['deleted']} 条")
        if total['skipped'] > 0:
            parts.append(f"跳过 {total['skipped']} 条")
        if total['parse_failed'] > 0:
            parts.append(f"{total['parse_failed']} 条解析失败")
        if total['errors'] > 0:
            parts.append(f"{total['errors']} 个文档出错")
        msg = "：".join(parts[:1]) + "，".join(parts[1:]) if len(parts) > 1 else parts[0]
        return {"message": msg, **total}
    except Exception as e:
        logger.error(f"反向同步接口错误: {e}", exc_info=True)
        raise HTTPException(status_code=502, detail=f"反向同步失败: {str(e)}")


@router.post("/import")
async def import_qa(
    file: UploadFile = File(...),
    user: User = Depends(require_kb_admin),
    db: AsyncSession = Depends(get_db),
):
    """Excel/CSV批量导入Q&A"""
    content = await file.read()
    filename = file.filename or "import.csv"

    count = 0
    skipped = 0  # 重复跳过计数
    try:
        # 解析所有行为 (question, answer) 列表
        rows_to_import: list[tuple[str, str]] = []
        if filename.endswith('.csv'):
            # CSV导入
            text = content.decode("utf-8-sig")
            lines = text.strip().split("\n")
            for i, line in enumerate(lines):
                if i == 0:
                    continue  # 跳过表头
                parts = _parse_csv_line(line)
                if len(parts) >= 2 and parts[0].strip() and parts[1].strip():
                    rows_to_import.append((parts[0].strip(), parts[1].strip()))
        else:
            # Excel导入 (xlsx/xls)
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # 跳过表头
                if len(row) >= 2 and row[0] and row[1]:
                    rows_to_import.append((str(row[0]).strip(), str(row[1]).strip()))
            wb.close()

        # 逐条重复检测后插入（FR-14）
        for question, answer in rows_to_import:
            duplicate = await QADuplicateDetector.check(
                db, question, user.active_team_id
            )
            if duplicate:
                skipped += 1
                logger.info(f"导入跳过重复QA: '{question[:50]}' (匹配: {duplicate.match_type})")
                continue
            qa = QAMeta(
                id=str(uuid.uuid4()),
                team_id=user.active_team_id,
                question=question,
                answer=answer,
                question_summary=question[:200],
                answer_summary=answer[:200],
                edited_by=user.id,
            )
            db.add(qa)
            count += 1

        await db.flush()

    except Exception as e:
        logger.error(f"QA导入失败: {e}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"导入失败: {str(e)}")

    msg = f"成功导入 {count} 条问答"
    if skipped > 0:
        msg += f"，跳过 {skipped} 条重复"
    return {"message": msg, "count": count, "skipped": skipped}


def _parse_csv_line(line: str) -> list:
    """简单CSV行解析，支持双引号包裹"""
    result = []
    current = ""
    in_quotes = False
    for char in line:
        if char == '"':
            in_quotes = not in_quotes
        elif char == ',' and not in_quotes:
            result.append(current)
            current = ""
        else:
            current += char
    result.append(current)
    return result


@router.get("/template")
async def download_template():
    """下载导入模板"""
    content = "问题,答案\n请输入问题,请输入答案\n产品保修期是多久？,标准保修期为一年\n如何联系客服？,可拨打400-xxx-xxxx或发送邮件至support@example.com\n"
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=qa_template.csv"},
    )


@router.get("/{qa_id}/versions")
async def get_qa_versions(
    qa_id: str,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取Q&A版本历史"""
    result = await db.execute(select(QAMeta).where(QAMeta.id == qa_id))
    qa = result.scalar_one_or_none()
    if not qa:
        raise HTTPException(status_code=404, detail="Q&A不存在")
    return {"current_version": qa.version, "history": [{"version": qa.version, "updated_at": qa.updated_at.isoformat()}]}
